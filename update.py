# run this code to generate an updated graph of screentime usage

from openpyxl import load_workbook
from re import search
from math import floor
from datetime import date, timedelta

path = "screentime_tracker/HistoryReport.xlsx"
# wb is short for workbook
wb = load_workbook(path)

def sheetDates(sheet_dates, sheet_names):
    # creates a tupled list of all sheets generated from Samsung app in the form (date, name), where
    # date comes from their sheet names and name is the sheet name
    
    sheet_names = (name for name in sheet_names)
    for name in sheet_names:
        # having trouble using backreferences, just repeating the same group instead
        if search("^\d{2}-\d{2}-\d{4}_\d{2}-\d{2}-\d{2}$", name):
            # make sure all sheets have been generated on Fridays, if not throw an Exception
            if date(int(name[6:10]), int(name[3:5]), int(name[0:2])).weekday() != 4:
                raise Exception("Not all sheets were generated on a Friday")
            # append date object from sheet name with the sheet name. Date objects are in the form y,m,d
            sheet_dates.append((date(int(name[6:10]), int(name[3:5]), int(name[0:2])), name))

def appNames():
    # returns a tuple containing all the app names present as column headers in the Data sheet
    # sets wb.active to the Data sheet

    wb.active = wb.get_sheet_by_name("Data")
    ws = wb.active
    app_names_defined = wb.defined_names['app_names']
    # dests is a tuple generator of (worksheet title, cell range)
    dests = app_names_defined.destinations
    _, coord = next(dests)
    # Strangely enough, ws[cell_range_str] returns a single tuple with all the cells representing app names
    # inside the first inner tuple
    return ws[coord][0]

# SHOULD NOT NEED
def intToExcelCol(num):
    if num <= 26:
        return str(chr(num+64))
    first = chr(floor(num / 26) + 64)
    second = chr(num % 27 + 65)
    return str(first) + str(second)

def main():
    # get list of (date, name) tuples for each sheet
    sheet_dates = []
    sheetDates(sheet_dates, wb.sheetnames)
    # sort dates from the earliest to the most recent
    sheet_dates.sort(key=lambda tup: tup[1])

    '''
    # get the most recent date recorded in Data

    # the following string represents the range of cells representing the most recent date in the Data sheet
    cell_range_str = "B" + str(ws.max_row) + ":" + "D" + str(ws.max_row)
    # convert to a date. Strangely enough, ws[cell_range_str] returns a single tuple with a triple tuple
    # inside that has the 3 cell values corresponding to the date
    y, m, d = ws[cell_range_str][0]
    most_recent_data_date = date(y.internal_value, m.internal_value, d.internal_value)

    get_later_date = (date[0] for date in sheet_dates)
    i = -1
    while next(get_later_date) > most_recent_data_date:
        i = i+1
    # delete sheets from sheet_dates that are from dates prior to the most recent date in Data, i.e.
    # their values have already been added
    if i >= 0:
        del sheet_dates[i:]
    '''
    # MOST OF THE ABOVE CODE CAN BE REMOVED IF I JUST ADD A FEATURE AT THE END OF THIS CODE TO REMOVE SHEETS
    # WITH DATA THAT HAS ALREADY BEEN ADDED
    
    # get a tuple with all the app cells in the Data sheet
    app_names_in_data = appNames()
    #for app in app_names_in_data:
    #    print(app.internal_value)
    wb.active = wb.get_sheet_by_name("Data")
    ws = wb.active
    # first empty row that will be populated in Data sheet
    top_row = ws.max_row + 1
    time_info_defined = wb.defined_names['time_info']
    # dests is a tuple generator of (worksheet title, cell range)
    dests = time_info_defined.destinations
    _, coord = next(dests)
    # assuming order of time_info cells is Season, DOTW, Date but not assuming their coordinates in the sheet
    time_info_cells = ws[coord][0]
    season = "Summer"
    # how can I take in input?
    #season = input("What season is it? If you do not enter one, the last season in the table will be used.")
    #if not season:
    #    season = 
    days = ["Sa", "Su", "M", "Tu", "W", "Th", "F"]
    for sheet in sheet_dates:
        # start by adding the time info
        wb.active = wb.get_sheet_by_name("Data")
        ws = wb.active
        for cell in time_info_cells:
            if cell.value == "Season":
                for i in range(0,7):
                    ws.cell(column=cell.column, row=top_row+i, value=season)
            if cell.value == "DOTW":
                for i, day in zip(range(0,7), days):
                    ws.cell(column=cell.column, row=top_row+i, value=day)
            if cell.value == "Date":
                for i in range(0,7):
                    ws.cell(column=cell.column, row=top_row+i, value=sheet[0]-timedelta(days=6-i))

    wb.save(path)

    '''
        # go to the least recent sheet of data to be added first
        wb.active = wb.get_sheet_by_name(sheet[1])
        ws = wb.active
        # max_row row value is a total row
        get_cols = ws.iter_cols(min_row=0, max_row=ws.max_row-1, values_only=True)
        # dealing with column of apps
        # skip over first cell, which is blank
        apps = next(get_cols)[1:]
        # change over to Data sheet to write stuff in
        wb.active = wb.get_sheet_by_name("Data")
        ws = wb.active
        for app in apps:
           if app not in app_names_in_data:
               print(intToExcelCol(ws.max_column+1) + "1")
               ws[intToExcelCol(ws.max_column+1) + "1"] = app'''
    '''
    TODO:
    -add data from app-generated sheets to Data table column by column, adding new columns when necessary
    -format a Total column to capture total time from all app columns
    -delete sheets with added data

    #remove sheets from sheet_dates
    '''
if __name__ == "__main__":
    main()